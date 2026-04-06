import httpx
import openpyxl
from fastapi import APIRouter, Request
from datetime import date, datetime, timedelta
from io import BytesIO
from app.config import TELEGRAM_TOKEN
from app.db.session import SessionLocal
from app.db import repo
from app.telegram.keyboards import category_keyboard

router = APIRouter()
TELEGRAM_API = f"https://api.telegram.org/bot{TELEGRAM_TOKEN}"

# in-memory state: {chat_id: {"step": str, "data": dict}}
user_state: dict = {}


# ── Telegram helpers ───────────────────────────────────────────────────────

async def send_message(chat_id: int, text: str, reply_markup=None):
    payload = {"chat_id": chat_id, "text": text, "parse_mode": "Markdown"}
    if reply_markup:
        payload["reply_markup"] = reply_markup
    async with httpx.AsyncClient() as client:
        await client.post(f"{TELEGRAM_API}/sendMessage", json=payload)


async def answer_callback_query(callback_query_id: str):
    async with httpx.AsyncClient() as client:
        await client.post(f"{TELEGRAM_API}/answerCallbackQuery",
                          json={"callback_query_id": callback_query_id})


def parse_amount(text: str) -> int | None:
    """Accepts formats like '25k' → 25000, '1.5k' → 1500."""
    try:
        text = text.strip().lower()
        if text.endswith("k"):
            return int(float(text[:-1]) * 1000)
        return int(text)
    except ValueError:
        return None


def format_summary(rows: list[tuple[str, int]], title: str) -> str:
    if not rows:
        return f"📊 {title}\n\nNo expenses recorded."
    width = max(len(name) for name, _ in rows)
    lines = [f"📊 *{title}*\n"]
    for name, total in rows:
        lines.append(f"{name:<{width}}   {total:,}")
    lines.append("─" * (width + 10))
    lines.append(f"{'Total':<{width}}   {sum(t for _, t in rows):,}")
    return "\n".join(lines)


def get_category_keyboard_for(db, user_id: int):
    cats = repo.get_categories(db, user_id)
    return category_keyboard([c.name for c in cats])


def inline_category_keyboard(categories):
    """Inline keyboard for mapping category selection."""
    buttons = [[{"text": c.name, "callback_data": f"map_category:{c.id}"}] for c in categories]
    return {"inline_keyboard": buttons}


# ── Signup flow ────────────────────────────────────────────────────────────

async def signup(chat_id: int, text: str, db) -> bool:
    """Returns True when signup is complete."""
    state = user_state.get(chat_id, {"step": "first_name", "data": {}})
    step = state["step"]

    if step == "first_name":
        await send_message(chat_id, "Welcome! Let's get you set up.\n\nEnter your first name:")
        state["step"] = "last_name"

    elif step == "last_name":
        state["data"]["first_name"] = text
        await send_message(chat_id, "Enter your last name (or `-` to skip):")
        state["step"] = "username"

    elif step == "username":
        state["data"]["last_name"] = text if text != "-" else None
        await send_message(chat_id, "Choose a username (must be unique):")
        state["step"] = "budget"

    elif step == "budget":
        username = text.strip()
        if repo.username_exists(db, username):
            await send_message(chat_id, "❌ That username is already taken. Please choose another:")
            user_state[chat_id] = state
            return False
        state["data"]["username"] = username
        await send_message(chat_id, "Enter your monthly budget (or `-` to skip):")
        state["step"] = "confirm_budget"

    elif step == "confirm_budget":
        budget = None
        if text != "-":
            budget = parse_amount(text)
            if budget is None:
                await send_message(chat_id, "❌ Invalid number. Enter your budget or `-` to skip:")
                user_state[chat_id] = state
                return False

        data = state["data"]
        user = repo.create_user(
            db,
            chat_id=chat_id,
            first_name=data.get("first_name"),
            last_name=data.get("last_name"),
            username=data["username"],
            budget=budget,
        )
        repo.seed_default_categories(db, user.id)
        user_state.pop(chat_id, None)

        kb = get_category_keyboard_for(db, chat_id)
        await send_message(chat_id, "✅ All set! Choose a category to log an expense:", reply_markup=kb)
        return True

    user_state[chat_id] = state
    return False


# ── Expense flow ───────────────────────────────────────────────────────────

async def expense_input(chat_id: int, text: str, db):
    state = user_state.get(chat_id)
    kb = get_category_keyboard_for(db, chat_id)

    if not state or state.get("step") == "awaiting_category":
        category = repo.get_category_by_name(db, chat_id, text)
        if not category:
            await send_message(chat_id, "❌ Unknown category. Please choose from the keyboard:", reply_markup=kb)
            return
        user_state[chat_id] = {"step": "awaiting_amount", "category_id": category.id, "category_name": category.name}
        await send_message(chat_id, f"*{text}* selected. Enter amount (or /cancel):")
        return

    if state.get("step") == "awaiting_amount":
        if text == "/cancel":
            user_state.pop(chat_id, None)
            await send_message(chat_id, "Cancelled. Choose a category:", reply_markup=kb)
            return

        amount = parse_amount(text)
        if amount is None or amount <= 0:
            await send_message(chat_id, "❌ Invalid amount. Enter a positive number (e.g. 500 or 25k), or /cancel:")
            return

        repo.create_expense(db, user_id=chat_id, category_id=state["category_id"], amount=amount)
        user_state.pop(chat_id, None)
        await send_message(
            chat_id,
            f"✅ *{state['category_name']}* — {amount:,} saved",
            reply_markup=kb,
        )


# ── Summary commands ───────────────────────────────────────────────────────

async def handle_summary(chat_id: int, period: str, db):
    today = date.today()
    if period == "day":
        start, end = today, today
        title = f"Today — {today.strftime('%b %d')}"
    elif period == "week":
        start = today - timedelta(days=today.weekday())
        end = today
        title = f"This Week ({start.strftime('%b %d')} – {end.strftime('%b %d')})"
    else:  # month
        start = today.replace(day=1)
        end = today
        title = today.strftime("%B %Y")

    rows = repo.get_expenses_summary(db, chat_id, start, end)
    await send_message(chat_id, format_summary(rows, title))


# ── Category management commands ───────────────────────────────────────────

async def handle_add_category(chat_id: int, text: str, db):
    parts = text.split(maxsplit=1)
    if len(parts) < 2 or not parts[1].strip():
        await send_message(chat_id, "Usage: `/add_category <name>`")
        return
    name = parts[1].strip()
    result = repo.add_category(db, chat_id, name)
    if result is None:
        await send_message(chat_id, f"❌ Category *{name}* already exists.")
    else:
        kb = get_category_keyboard_for(db, chat_id)
        await send_message(chat_id, f"✅ Category *{name}* added.", reply_markup=kb)


async def handle_remove_category(chat_id: int, text: str, db):
    parts = text.split(maxsplit=1)
    if len(parts) < 2 or not parts[1].strip():
        await send_message(chat_id, "Usage: `/remove_category <name>`")
        return
    name = parts[1].strip()
    deleted = repo.remove_category(db, chat_id, name)
    if not deleted:
        await send_message(chat_id, f"❌ Category *{name}* not found.")
    else:
        kb = get_category_keyboard_for(db, chat_id)
        await send_message(chat_id, f"✅ Category *{name}* removed.", reply_markup=kb)


# ── Service mapping commands ────────────────────────────────────────────────

async def handle_add_mapping(chat_id: int, db):
    user_state[chat_id] = {"step": "awaiting_mapping_keyword"}
    await send_message(chat_id, "Enter the keyword(s) for the service (e.g. `baraka market`), or /cancel:")


async def handle_list_mappings(chat_id: int, db):
    mappings = repo.get_service_mappings(db, chat_id)
    if not mappings:
        await send_message(chat_id, "No mappings yet. Use /add\\_mapping to create one.")
        return
    lines = ["*Your mappings:*\n"]
    for m in mappings:
        lines.append(f"• `{m.keyword}` → {m.category.name}")
    await send_message(chat_id, "\n".join(lines))


async def handle_remove_mapping(chat_id: int, text: str, db):
    parts = text.split(maxsplit=1)
    if len(parts) < 2 or not parts[1].strip():
        await send_message(chat_id, "Usage: `/remove_mapping <keyword>`")
        return
    keyword = parts[1].strip()
    deleted = repo.remove_service_mapping(db, chat_id, keyword)
    if not deleted:
        await send_message(chat_id, f"❌ Mapping `{keyword}` not found.")
    else:
        await send_message(chat_id, f"✅ Mapping `{keyword}` removed.")


async def handle_mapping_keyword(chat_id: int, text: str, db):
    """Called when user has typed a keyword during /add_mapping flow."""
    keyword = text.strip().lower()
    if not keyword:
        await send_message(chat_id, "❌ Keyword can't be empty. Try again:")
        return
    user_state[chat_id] = {"step": "awaiting_mapping_category", "keyword": keyword}
    cats = repo.get_categories(db, chat_id)
    await send_message(chat_id, f"Keyword: `{keyword}`\n\nChoose a category:",
                       reply_markup=inline_category_keyboard(cats))


async def handle_callback_query(cq: dict, db):
    """Handles inline keyboard button presses."""
    callback_query_id = cq["id"]
    chat_id = cq["from"]["id"]
    data = cq.get("data", "")

    await answer_callback_query(callback_query_id)

    if data.startswith("map_category:"):
        category_id = int(data.split(":")[1])
        state = user_state.get(chat_id, {})
        keyword = state.get("keyword")
        if not keyword:
            await send_message(chat_id, "❌ Session expired. Please start /add\\_mapping again.")
            return

        result = repo.add_service_mapping(db, chat_id, keyword, category_id)
        user_state.pop(chat_id, None)
        if result is None:
            await send_message(chat_id, f"❌ Mapping for `{keyword}` already exists.")
        else:
            await send_message(chat_id, f"✅ `{keyword}` → *{result.category.name}*")


# ── Click xlsx import ───────────────────────────────────────────────────────

async def handle_xlsx_import(chat_id: int, file_id: str, db):
    async with httpx.AsyncClient() as client:
        r = await client.get(f"{TELEGRAM_API}/getFile", params={"file_id": file_id})
        file_path = r.json()["result"]["file_path"]
        file_r = await client.get(f"https://api.telegram.org/file/bot{TELEGRAM_TOKEN}/{file_path}")
        content = file_r.content

    wb = openpyxl.load_workbook(BytesIO(content))
    ws = wb.active

    # Map header names to column indices
    headers = {str(cell.value).strip(): idx for idx, cell in enumerate(ws[1])}
    required = {"Сумма", "Время", "Карта", "Сервис", "Статус платежа"}
    if not required.issubset(headers):
        await send_message(chat_id, "❌ Unrecognized file format. Expected Click export.")
        return

    imported = duplicates = unmatched = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        status = str(row[headers["Статус платежа"]]).strip()
        if status != "Успешно проведен":
            continue

        time_val = row[headers["Время"]]
        card = str(row[headers["Карта"]]).strip()

        # Normalize time to string for import_ref
        if isinstance(time_val, datetime):
            time_str = time_val.strftime("%d.%m.%Y %H:%M:%S")
            expense_date = time_val.date()
        else:
            time_str = str(time_val).strip()
            try:
                expense_date = datetime.strptime(time_str, "%d.%m.%Y %H:%M:%S").date()
            except ValueError:
                unmatched += 1
                continue

        import_ref = f"{time_str}|{card}"
        if repo.import_ref_exists(db, chat_id, import_ref):
            duplicates += 1
            continue

        service = str(row[headers["Сервис"]]).strip()
        category = repo.get_category_for_service(db, chat_id, service)
        if not category:
            unmatched += 1
            continue

        amount = round(float(row[headers["Сумма"]]))
        if amount <= 0:
            continue

        repo.create_imported_expense(db, chat_id, category.id, amount, expense_date, import_ref)
        imported += 1

    await send_message(
        chat_id,
        f"✅ Imported: *{imported}* | ⏭ Duplicates: *{duplicates}* | ❓ Unmatched: *{unmatched}*"
    )


# ── Webhook entry point ────────────────────────────────────────────────────

@router.post("/webhook")
async def telegram_webhook(req: Request):
    data = await req.json()

    db = SessionLocal()
    try:
        # Handle inline keyboard button presses
        if "callback_query" in data:
            await handle_callback_query(data["callback_query"], db)
            return {"ok": True}

        if "message" not in data:
            return {"ok": True}

        msg = data["message"]
        chat_id = msg["chat"]["id"]
        text = msg.get("text", "").strip()

        user = repo.get_user(db, chat_id)

        if not user:
            if text == "/start" and chat_id not in user_state:
                await send_message(chat_id,
                    "👋 Welcome to *FinTrack* — a minimalist expense tracker.\n\n"
                    "Tap a category, enter an amount, done. "
                    "Use /day, /week, /month to see your spending.\n\n"
                    "Let's set up your account first."
                )
            await signup(chat_id, text, db)
            return {"ok": True}

        # Handle document upload (xlsx import)
        doc = msg.get("document")
        if doc:
            state = user_state.get(chat_id, {})
            if state.get("step") == "awaiting_import_file":
                user_state.pop(chat_id, None)
                xlsx_mime = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                if doc.get("mime_type") == xlsx_mime:
                    await handle_xlsx_import(chat_id, doc["file_id"], db)
                else:
                    await send_message(chat_id, "❌ Please send an .xlsx file.")
            return {"ok": True}

        # Handle awaiting_mapping_keyword state before command routing
        state = user_state.get(chat_id, {})
        if state.get("step") == "awaiting_mapping_keyword" and not text.startswith("/"):
            await handle_mapping_keyword(chat_id, text, db)
            return {"ok": True}

        # Commands
        if text == "/day":
            await handle_summary(chat_id, "day", db)
        elif text == "/week":
            await handle_summary(chat_id, "week", db)
        elif text == "/month":
            await handle_summary(chat_id, "month", db)
        elif text == "/start":
            kb = get_category_keyboard_for(db, chat_id)
            await send_message(chat_id, f"Hey, {user.first_name}! Choose a category:", reply_markup=kb)
        elif text == "/cancel":
            user_state.pop(chat_id, None)
            kb = get_category_keyboard_for(db, chat_id)
            await send_message(chat_id, "Cancelled. Choose a category:", reply_markup=kb)
        elif text == "/import":
            user_state[chat_id] = {"step": "awaiting_import_file"}
            await send_message(chat_id, "Send your Click `.xlsx` file (or /cancel):")
        elif text == "/add_mapping":
            await handle_add_mapping(chat_id, db)
        elif text == "/list_mappings":
            await handle_list_mappings(chat_id, db)
        elif text.startswith("/remove_mapping"):
            await handle_remove_mapping(chat_id, text, db)
        elif text.startswith("/add_category"):
            await handle_add_category(chat_id, text, db)
        elif text.startswith("/remove_category"):
            await handle_remove_category(chat_id, text, db)
        elif text.startswith("/"):
            await send_message(chat_id, (
                "Unknown command. Available commands:\n\n"
                "/day — today's expenses\n"
                "/week — this week's expenses\n"
                "/month — this month's expenses\n"
                "/import — import Click .xlsx file\n"
                "/add\\_mapping — map a service keyword to a category\n"
                "/list\\_mappings — show all keyword mappings\n"
                "/remove\\_mapping <keyword> — remove a mapping\n"
                "/add\\_category <name> — add a category\n"
                "/remove\\_category <name> — remove a category\n"
                "/cancel — cancel current input"
            ))
        else:
            await expense_input(chat_id, text, db)
    finally:
        db.close()

    return {"ok": True}
